"""
导出 Excel 格式的 TikTok 评论数据 - Vercel Serverless 函数
"""

from flask import Flask, request, jsonify, send_file
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)


@app.route('/', methods=['POST'])
@app.route('/api/export/excel', methods=['POST'])
def handler():
    """导出Excel格式的评论数据"""
    try:
        data = request.get_json()
        videos = data.get('videos', [])

        if not videos:
            return jsonify({
                "success": False,
                "error": "没有视频数据可导出"
            }), 400

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "TikTok评论汇总"

        # 定义样式
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 设置列宽
        col_widths = [20]  # 第一列（标签列）
        for video in videos:
            col_widths.extend([25, 40, 10, 10])  # 每个视频4列

        for i, width in enumerate(col_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # 写入表头 - 第一列
        header_cell = ws.cell(row=1, column=1, value="视频信息")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.border = border
        header_cell.alignment = alignment

        # 写入视频标题 - 每个视频标题占据4列
        for i, video in enumerate(videos, 1):
            # 起始列（每个视频占4列）
            start_col = 2 + ((i - 1) * 4)
            end_col = start_col + 3

            # 合并单元格
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            # 写入视频标题
            video_title = f"视频 {i}"
            cell = ws.cell(row=1, column=start_col, value=video_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment

            # 为合并的其他单元格设置边框和样式
            for col in range(start_col + 1, end_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.border = border
                cell.fill = header_fill

        # 写入视频信息 - 每个视频信息合并4列
        current_row = 2

        # 第一列：标签
        label_cell = ws.cell(row=current_row, column=1, value="视频信息")
        label_cell.fill = info_fill
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.font = Font(bold=True)

        # 为每个视频写一行综合信息
        for video_index, video in enumerate(videos):
            # 起始列（每个视频占4列）
            start_col = 2 + (video_index * 4)
            end_col = start_col + 3

            # 合并单元格
            ws.merge_cells(start_row=current_row, start_column=start_col,
                          end_row=current_row, end_column=end_col)

            # 写入视频信息
            video_info = f"视频 {video_index + 1}\n{video.get('url', 'N/A')}\n{video.get('video_id', 'N/A')}\n{video.get('total_comments', 0)}条"
            cell = ws.cell(row=current_row, column=start_col, value=video_info)
            cell.fill = info_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=False, size=10)

            # 为合并的其他单元格设置边框
            for col in range(start_col + 1, end_col + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.fill = info_fill

        # 设置行高
        ws.row_dimensions[current_row].height = 80
        current_row += 1

        # 写入评论列表标题 - 使用更细的列结构
        comments_header_row = current_row

        # 第一列：评论列表标签
        comments_header_cell = ws.cell(row=comments_header_row, column=1, value="评论列表")
        comments_header_cell.fill = header_fill
        comments_header_cell.border = border
        comments_header_cell.alignment = alignment
        comments_header_cell.font = header_font

        # 为每个视频创建4个子列：评论者、评论内容、点赞数、评论数
        col_index = 2
        for video_index, video in enumerate(videos):
            # 子列标题
            sub_headers = [
                (col_index, f"视频 {video_index + 1}-评论者"),
                (col_index + 1, f"视频 {video_index + 1}-评论内容"),
                (col_index + 2, f"视频 {video_index + 1}-点赞数"),
                (col_index + 3, f"视频 {video_index + 1}-评论数")
            ]

            for sub_col, sub_title in sub_headers:
                cell = ws.cell(row=comments_header_row, column=sub_col)
                cell.value = sub_title
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
                cell.font = header_font

                # 设置子列宽度
                col_letter = get_column_letter(sub_col)
                if "评论内容" in sub_title:
                    ws.column_dimensions[col_letter].width = 40
                elif "评论者" in sub_title:
                    ws.column_dimensions[col_letter].width = 20
                elif "点赞数" in sub_title:
                    ws.column_dimensions[col_letter].width = 10
                elif "评论数" in sub_title:
                    ws.column_dimensions[col_letter].width = 10

            col_index += 4

        current_row += 1

        # 写入评论数据
        max_comments = max(len(video.get('comments', [])) for video in videos) if videos else 0

        for comment_index in range(max_comments):
            # 标签列
            label_cell = ws.cell(row=current_row, column=1, value=f"评论 {comment_index + 1}")
            label_cell.border = border
            label_cell.alignment = alignment

            # 每个视频的对应评论（4列一组）
            col_index = 2
            for video in videos:
                comments = video.get('comments', [])

                if comment_index < len(comments):
                    comment = comments[comment_index]
                    author = comment.get('author', {}).get('nickname', '未知用户')
                    text = comment.get('text', '')
                    likes = comment.get('likes', 0)
                    reply_count = comment.get('reply_count', 0)

                    # 评论者
                    author_cell = ws.cell(row=current_row, column=col_index, value=author)
                    author_cell.border = border
                    author_cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 评论内容
                    text_cell = ws.cell(row=current_row, column=col_index + 1, value=text)
                    text_cell.border = border
                    text_cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 点赞
                    likes_cell = ws.cell(row=current_row, column=col_index + 2, value=likes)
                    likes_cell.border = border
                    likes_cell.alignment = Alignment(horizontal='center', vertical='top')

                    # 评论（回复数）
                    reply_cell = ws.cell(row=current_row, column=col_index + 3, value=reply_count)
                    reply_cell.border = border
                    reply_cell.alignment = Alignment(horizontal='center', vertical='top')
                else:
                    # 空白单元格
                    for i in range(4):
                        cell = ws.cell(row=current_row, column=col_index + i, value="")
                        cell.border = border
                        cell.alignment = Alignment(vertical='top')

                col_index += 4

            current_row += 1

        # 冻结前两行
        ws.freeze_panes = "A3"

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'tiktok_comments_batch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Excel导出失败: {str(e)}"
        }), 500
