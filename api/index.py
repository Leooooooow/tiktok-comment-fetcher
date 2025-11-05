"""
TikTok 评论获取器 - Vercel Serverless 版本
"""

from flask import Flask, render_template, request, jsonify, send_file
import requests
import re
import json
import csv
import os
import sys
from io import StringIO, BytesIO
from datetime import datetime
from typing import Dict, Any, List
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 获取项目根目录
base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, base_dir)

# 设置模板和静态文件路径
template_dir = os.path.join(base_dir, 'templates')
static_dir = os.path.join(base_dir, 'static')

app = Flask(__name__,
            template_folder=template_dir,
            static_folder=static_dir,
            static_url_path='/static')

# API 配置 - 从环境变量读取（更安全）
API_KEY = os.environ.get('TIKHUB_API_KEY', "yY08aG9D6Gt45xNfyVW/s2oZ0kAkzYzcqMxwkGb27TJErnoTdfwowAWLEA==")
BASE_URL = "https://api.tikhub.io"


def extract_video_id(url: str) -> str:
    """从 TikTok URL 中提取视频 ID"""
    patterns = [
        r'/video/(\d+)',
        r'/v/(\d+)',
        r'tiktok\.com/.*?/video/(\d+)',
        r'vm\.tiktok\.com/(\w+)',
        r'vt\.tiktok\.com/(\w+)'
    ]

    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)

    return None


def fetch_comments_app_v3(aweme_id: str, cursor: int = 0, count: int = 30) -> Dict[Any, Any]:
    """使用 APP V3 接口获取评论"""
    endpoint = f"{BASE_URL}/api/v1/tiktok/app/v3/fetch_video_comments"

    headers = {
        "Authorization": f"Bearer {API_KEY}",
        "Content-Type": "application/json"
    }

    params = {
        "aweme_id": aweme_id,
        "cursor": cursor,
        "count": count
    }

    try:
        response = requests.get(endpoint, params=params, headers=headers, timeout=30)

        if response.status_code == 200:
            return response.json()
        else:
            return {
                "error": f"HTTP {response.status_code}",
                "message": response.text[:200]
            }
    except requests.exceptions.Timeout:
        return {"error": "请求超时，请重试"}
    except requests.exceptions.ConnectionError:
        return {"error": "网络连接失败，请检查网络"}
    except Exception as e:
        return {"error": f"请求失败: {str(e)}"}


def fetch_all_comments(aweme_id: str) -> Dict[str, Any]:
    """获取视频的所有评论（自动翻页）"""
    all_comments = []
    cursor = 0
    count = 30
    has_more = True
    max_pages = 50  # 限制最大页数，防止 serverless 超时

    while has_more and len(all_comments) < 1500:  # 限制最大评论数
        result = fetch_comments_app_v3(aweme_id, cursor=cursor, count=count)

        if "error" in result:
            return {
                "success": False,
                "error": result.get("error", "未知错误"),
                "message": result.get("message", ""),
                "total_comments": 0,
                "comments": []
            }

        data = result.get("data", {})
        comments = data.get("comments", [])

        if not comments:
            break

        all_comments.extend(comments)

        has_more = data.get("has_more", False)
        cursor = data.get("cursor", 0)

        if not has_more or cursor == 0:
            break

    return {
        "success": True,
        "aweme_id": aweme_id,
        "total_comments": len(all_comments),
        "comments": all_comments
    }


def format_comment(comment: Dict) -> Dict:
    """格式化评论数据"""
    user = comment.get("user", {})
    avatar_data = user.get("avatar_thumb", {})
    avatar_urls = avatar_data.get("url_list", [])

    return {
        "id": comment.get("cid", ""),
        "text": comment.get("text", ""),
        "author": {
            "uid": user.get("uid", ""),
            "nickname": user.get("nickname", "Unknown"),
            "username": user.get("unique_id", ""),
            "avatar": avatar_urls[0] if avatar_urls else "",
            "signature": user.get("signature", "")
        },
        "likes": comment.get("digg_count", 0),
        "create_time": comment.get("create_time", 0),
        "create_time_formatted": datetime.fromtimestamp(
            comment.get("create_time", 0)
        ).strftime("%Y-%m-%d %H:%M:%S"),
        "reply_count": comment.get("reply_comment_total", 0),
        "status": comment.get("status", 0)
    }


@app.route('/')
def index():
    """主页"""
    try:
        return render_template('index.html')
    except Exception as e:
        # 调试信息
        import traceback
        error_details = {
            "error": str(e),
            "template_folder": app.template_folder,
            "static_folder": app.static_folder,
            "base_dir": base_dir,
            "template_exists": os.path.exists(os.path.join(template_dir, 'index.html')),
            "traceback": traceback.format_exc()
        }
        return jsonify(error_details), 500


@app.route('/api/fetch-comments', methods=['POST'])
def fetch_comments_route():
    """获取评论的 API 端点"""
    try:
        data = request.get_json()
        url = data.get('url', '').strip()

        if not url:
            return jsonify({
                "success": False,
                "error": "请输入 TikTok 视频 URL"
            }), 400

        video_id = extract_video_id(url)

        if not video_id:
            return jsonify({
                "success": False,
                "error": "无法从 URL 中提取视频 ID，请检查 URL 格式"
            }), 400

        result = fetch_all_comments(video_id)

        if not result["success"]:
            return jsonify(result), 500

        formatted_comments = [format_comment(c) for c in result["comments"]]

        return jsonify({
            "success": True,
            "video_id": video_id,
            "total": result["total_comments"],
            "comments": formatted_comments
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"服务器错误: {str(e)}"
        }), 500


@app.route('/api/export/<format>', methods=['POST'])
def export_comments(format):
    """导出评论数据"""
    try:
        data = request.get_json()
        comments = data.get('comments', [])
        video_id = data.get('video_id', 'unknown')

        if format == 'json':
            output = BytesIO()
            output.write(json.dumps(comments, ensure_ascii=False, indent=2).encode('utf-8'))
            output.seek(0)

            return send_file(
                output,
                mimetype='application/json',
                as_attachment=True,
                download_name=f'tiktok_comments_{video_id}.json'
            )

        elif format == 'csv':
            output = StringIO()
            writer = csv.writer(output)

            writer.writerow([
                'Comment ID', 'Author', 'Username', 'Comment Text',
                'Likes', 'Reply Count', 'Created Time'
            ])

            for comment in comments:
                writer.writerow([
                    comment.get('id', ''),
                    comment.get('author', {}).get('nickname', ''),
                    comment.get('author', {}).get('username', ''),
                    comment.get('text', ''),
                    comment.get('likes', 0),
                    comment.get('reply_count', 0),
                    comment.get('create_time_formatted', '')
                ])

            output.seek(0)
            output_bytes = BytesIO(output.getvalue().encode('utf-8-sig'))

            return send_file(
                output_bytes,
                mimetype='text/csv',
                as_attachment=True,
                download_name=f'tiktok_comments_{video_id}.csv'
            )

        else:
            return jsonify({
                "success": False,
                "error": "不支持的导出格式"
            }), 400

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"导出失败: {str(e)}"
        }), 500


@app.route('/api/fetch-comments-batch', methods=['POST'])
def fetch_comments_batch_route():
    """批量获取评论的 API 端点"""
    try:
        data = request.get_json()
        urls = data.get('urls', [])

        if not urls:
            return jsonify({
                "success": False,
                "error": "请输入至少一个 TikTok 视频 URL"
            }), 400

        # 去重和清理
        clean_urls = list(set(url.strip() for url in urls if url.strip()))

        if len(clean_urls) == 0:
            return jsonify({
                "success": False,
                "error": "没有有效的 TikTok 视频 URL"
            }), 400

        if len(clean_urls) > 10:
            return jsonify({
                "success": False,
                "error": "最多支持同时处理 10 个视频链接"
            }), 400

        # 处理每个视频
        results = []
        total_comments = 0
        successful_videos = 0

        def process_single_video(url):
            try:
                video_id = extract_video_id(url)
                if not video_id:
                    return {
                        "url": url,
                        "success": False,
                        "error": "无法从 URL 中提取视频 ID",
                        "video_id": None,
                        "total_comments": 0,
                        "comments": []
                    }

                result = fetch_all_comments(video_id)
                if not result["success"]:
                    return {
                        "url": url,
                        "success": False,
                        "error": result.get("error", "获取评论失败"),
                        "video_id": video_id,
                        "total_comments": 0,
                        "comments": []
                    }

                formatted_comments = [format_comment(c) for c in result["comments"]]
                return {
                    "url": url,
                    "success": True,
                    "error": None,
                    "video_id": video_id,
                    "total_comments": len(formatted_comments),
                    "comments": formatted_comments
                }
            except Exception as e:
                return {
                    "url": url,
                    "success": False,
                    "error": f"处理失败: {str(e)}",
                    "video_id": None,
                    "total_comments": 0,
                    "comments": []
                }

        # 使用线程池并发处理（最多5个线程）
        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_url = {executor.submit(process_single_video, url): url for url in clean_urls}

            for future in as_completed(future_to_url):
                result = future.result()
                results.append(result)

                if result["success"]:
                    successful_videos += 1
                    total_comments += result["total_comments"]

        # 保持原始URL顺序
        url_to_result = {r["url"]: r for r in results}
        ordered_results = [url_to_result[url] for url in clean_urls if url in url_to_result]

        return jsonify({
            "success": True,
            "total_videos": len(clean_urls),
            "successful_videos": successful_videos,
            "total_comments": total_comments,
            "videos": ordered_results
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"批量处理失败: {str(e)}"
        }), 500


@app.route('/api/export/excel', methods=['POST'])
def export_excel():
    """导出Excel格式的评论数据"""
    try:
        data = request.get_json()
        videos = data.get('videos', [])

        if not videos:
            return jsonify({
                "success": False,
                "error": "没有视频数据可导出"
            }), 400

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "TikTok评论汇总"

        # 定义样式
        header_font = Font(size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        info_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # 设置列宽
        col_widths = [20]  # 第一列（标签列）
        for video in videos:
            col_widths.extend([25, 40, 10, 10])  # 每个视频4列

        for i, width in enumerate(col_widths, 1):
            col_letter = get_column_letter(i)
            ws.column_dimensions[col_letter].width = width

        # 写入表头 - 第一列
        header_cell = ws.cell(row=1, column=1, value="视频信息")
        header_cell.font = header_font
        header_cell.fill = header_fill
        header_cell.border = border
        header_cell.alignment = alignment

        # 写入视频标题 - 每个视频标题占据4列
        for i, video in enumerate(videos, 1):
            # 起始列（每个视频占4列）
            start_col = 2 + ((i - 1) * 4)
            end_col = start_col + 3

            # 合并单元格
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)

            # 写入视频标题
            video_title = f"视频 {i}"
            cell = ws.cell(row=1, column=start_col, value=video_title)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment

            # 为合并的其他单元格设置边框和样式
            for col in range(start_col + 1, end_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.border = border
                cell.fill = header_fill

        # 写入视频信息 - 每个视频信息合并4列
        current_row = 2

        # 第一列：标签
        label_cell = ws.cell(row=current_row, column=1, value="视频信息")
        label_cell.fill = info_fill
        label_cell.border = border
        label_cell.alignment = Alignment(horizontal='center', vertical='center')
        label_cell.font = Font(bold=True)

        # 为每个视频写一行综合信息
        for video_index, video in enumerate(videos):
            # 起始列（每个视频占4列）
            start_col = 2 + (video_index * 4)
            end_col = start_col + 3

            # 合并单元格
            ws.merge_cells(start_row=current_row, start_column=start_col,
                          end_row=current_row, end_column=end_col)

            # 写入视频信息
            video_info = f"视频 {video_index + 1}\n{video.get('url', 'N/A')}\n{video.get('video_id', 'N/A')}\n{video.get('total_comments', 0)}条"
            cell = ws.cell(row=current_row, column=start_col, value=video_info)
            cell.fill = info_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.font = Font(bold=False, size=10)

            # 为合并的其他单元格设置边框
            for col in range(start_col + 1, end_col + 1):
                cell = ws.cell(row=current_row, column=col)
                cell.border = border
                cell.fill = info_fill

        # 设置行高
        ws.row_dimensions[current_row].height = 80
        current_row += 1

        # 写入评论列表标题 - 使用更细的列结构
        comments_header_row = current_row

        # 第一列：评论列表标签
        comments_header_cell = ws.cell(row=comments_header_row, column=1, value="评论列表")
        comments_header_cell.fill = header_fill
        comments_header_cell.border = border
        comments_header_cell.alignment = alignment
        comments_header_cell.font = header_font

        # 为每个视频创建4个子列：评论者、评论内容、点赞数、评论数
        col_index = 2
        for video_index, video in enumerate(videos):
            # 子列标题
            sub_headers = [
                (col_index, f"视频 {video_index + 1}-评论者"),
                (col_index + 1, f"视频 {video_index + 1}-评论内容"),
                (col_index + 2, f"视频 {video_index + 1}-点赞数"),
                (col_index + 3, f"视频 {video_index + 1}-评论数")
            ]

            for sub_col, sub_title in sub_headers:
                cell = ws.cell(row=comments_header_row, column=sub_col)
                cell.value = sub_title
                cell.fill = header_fill
                cell.border = border
                cell.alignment = alignment
                cell.font = header_font

                # 设置子列宽度
                col_letter = get_column_letter(sub_col)
                if "评论内容" in sub_title:
                    ws.column_dimensions[col_letter].width = 40
                elif "评论者" in sub_title:
                    ws.column_dimensions[col_letter].width = 20
                elif "点赞数" in sub_title:
                    ws.column_dimensions[col_letter].width = 10
                elif "评论数" in sub_title:
                    ws.column_dimensions[col_letter].width = 10

            col_index += 4

        current_row += 1

        # 写入评论数据
        max_comments = max(len(video.get('comments', [])) for video in videos) if videos else 0

        for comment_index in range(max_comments):
            # 标签列
            label_cell = ws.cell(row=current_row, column=1, value=f"评论 {comment_index + 1}")
            label_cell.border = border
            label_cell.alignment = alignment

            # 每个视频的对应评论（4列一组）
            col_index = 2
            for video in videos:
                comments = video.get('comments', [])

                if comment_index < len(comments):
                    comment = comments[comment_index]
                    author = comment.get('author', {}).get('nickname', '未知用户')
                    text = comment.get('text', '')
                    likes = comment.get('likes', 0)
                    reply_count = comment.get('reply_count', 0)

                    # 评论者
                    author_cell = ws.cell(row=current_row, column=col_index, value=author)
                    author_cell.border = border
                    author_cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 评论内容
                    text_cell = ws.cell(row=current_row, column=col_index + 1, value=text)
                    text_cell.border = border
                    text_cell.alignment = Alignment(vertical='top', wrap_text=True)

                    # 点赞
                    likes_cell = ws.cell(row=current_row, column=col_index + 2, value=likes)
                    likes_cell.border = border
                    likes_cell.alignment = Alignment(horizontal='center', vertical='top')

                    # 评论（回复数）
                    reply_cell = ws.cell(row=current_row, column=col_index + 3, value=reply_count)
                    reply_cell.border = border
                    reply_cell.alignment = Alignment(horizontal='center', vertical='top')
                else:
                    # 空白单元格
                    for i in range(4):
                        cell = ws.cell(row=current_row, column=col_index + i, value="")
                        cell.border = border
                        cell.alignment = Alignment(vertical='top')

                col_index += 4

            current_row += 1

        # 冻结前两行
        ws.freeze_panes = "A3"

        # 保存到内存
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'tiktok_comments_batch_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )

    except Exception as e:
        return jsonify({
            "success": False,
            "error": f"Excel导出失败: {str(e)}"
        }), 500


@app.route('/health')
def health():
    """健康检查端点"""
    return jsonify({
        "status": "healthy",
        "timestamp": datetime.now().isoformat()
    })


@app.route('/favicon.ico')
def favicon():
    """返回 favicon（避免 404 错误）"""
    # 返回一个简单的 SVG favicon
    svg = '''<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 100 100">
        <rect width="100" height="100" fill="#6366f1"/>
        <text x="50" y="70" font-size="60" text-anchor="middle" fill="white" font-family="Arial">T</text>
    </svg>'''
    return svg, 200, {'Content-Type': 'image/svg+xml'}


# Vercel serverless handler - 这是关键！
# Vercel 会查找名为 app 的 WSGI 应用
# 不需要额外的 handler 函数
